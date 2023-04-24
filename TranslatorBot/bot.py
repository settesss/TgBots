import random
import telebot
import config
from openpyxl import load_workbook

START_COMMAND = 'Добро пожаловать в Wooord Hunt!'

HELP_COMMAND = """
Привет! Я бот для работы с словарями Wooord Hunt.
Чтобы начать работу, напишите /start.
Для помощи, напишите /help."""

book = load_workbook(filename="D:/words.xlsx")
sheet = book['Лист1']
SHEET_ROWS_COUNT = sheet.max_row
row_number = None
question_word = None
answer_word = None

bot = telebot.TeleBot(config.TOKEN)


class UserState:
    def __init__(self):
        self.lesson = False
        self.incorrect_answers = []
        self.welcomed = False


user_state = UserState()


@bot.message_handler(commands=['start'])
def handle_start(message):
    user_markup = telebot.types.ReplyKeyboardMarkup(True, True)
    user_markup.row('Изучать слова 📖', 'Мои ошибки ⚠️')
    user_markup.row('Награды 🥇')
    user_markup.row('Помощь 🆘')
    user_markup.row('Записаться на урок 💎')
    if not user_state.welcomed:
        bot.send_message(message.chat.id, START_COMMAND, reply_markup=user_markup)
        user_state.welcomed = True
    else:
        bot.send_message(message.chat.id, '<b>Открываю меню...</b>', reply_markup=user_markup, parse_mode='HTML')
    bot.delete_message(message.chat.id, message.message_id)


@bot.message_handler(commands=['help'])
def handle_help(message):
    bot.send_message(message.chat.id, HELP_COMMAND)
    bot.delete_message(message.chat.id, message.message_id)


def ask_question(chat_id):
    global row_number, question_word, answer_word
    row_number = str(random.randrange(1, SHEET_ROWS_COUNT))
    question_word = sheet['A' + row_number].value
    answer_word = sheet['B' + row_number].value
    other_words = [sheet['B' + str(i)].value for i in range(1, SHEET_ROWS_COUNT) if
                   i != row_number]
    other_words_sampled = random.sample(other_words, 3)
    answer_cells = [answer_word] + other_words_sampled
    random.shuffle(answer_cells)
    answer_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    answer_markup.row(answer_cells[0], answer_cells[1])
    answer_markup.row(answer_cells[2], answer_cells[3])
    answer_markup.row("Я устал ⛔")
    bot.send_message(chat_id, question_word.title(), reply_markup=answer_markup)


@bot.message_handler(content_types=['text'])
def echo(message):
    if message.text.lower() == 'изучать слова 📖':
        user_state.lesson = True
        bot.send_message(message.chat.id, "Поехали!")
        ask_question(message.chat.id)
    elif message.text.lower() == 'мои ошибки ⚠️'.lower():
        incorrect_string = ', '.join(user_state.incorrect_answers)
        if len(user_state.incorrect_answers) != 0:
            bot.send_message(message.chat.id, incorrect_string)
        else:
            bot.send_message(message.chat.id, 'У Вас нет ошибок! Поздравляю!')
        handle_start(message)
    elif message.text.lower() == 'Помощь 🆘'.lower():
        handle_help(message)
    elif message.text.lower() == 'пока'.lower():
        bot.send_message(message.chat.id, "Только попробуй не прийти завтра на занятие XD")
    elif not answer_word and not user_state.lesson:
        bot.send_message(message.chat.id, "Введите 'изучать слова 📖' для начала изучения слов.")
        handle_start(message)
    elif message.text.lower() == answer_word:
        bot.send_message(message.chat.id, "Молодец!")
        ask_question(message.chat.id)
    elif message.text.lower() == 'Я устал ⛔'.lower():
        user_state.lesson = False
        handle_start(message)
    elif user_state.lesson:
        user_state.incorrect_answers.append(question_word)
        bot.send_message(message.chat.id, "Промах, я запишу это в ошибки!")
        ask_question(message.chat.id)
    print(f"{message.from_user.first_name}: {message.text}")


bot.polling()
